from __future__ import annotations

from pathlib import Path
from datetime import datetime

import pandas as pd


ROOT = Path(__file__).resolve().parents[1]

SUMMARY_PATH = ROOT / "data" / "official" / "processed" / "ir_derivatives_pricing_summary.csv"
SHOCK_PATH = ROOT / "data" / "official" / "processed" / "ir_swap_shock_table.csv"
LIFECYCLE_PATH = ROOT / "data" / "official" / "processed" / "ir_derivatives_model_lifecycle_register.csv"

EXPORT_DIR = ROOT / "excel_vba" / "exports"
TEMPLATE_DIR = ROOT / "excel_vba" / "templates"
REPORT_PATH = ROOT / "reports" / "excel_vba_revalidation_bridge.md"

CONTROL_EXPORT = EXPORT_DIR / "ir_swap_revalidation_control_panel.csv"
SHOCK_EXPORT = EXPORT_DIR / "ir_swap_revalidation_shock_table.csv"
LIFECYCLE_EXPORT = EXPORT_DIR / "ir_swap_revalidation_lifecycle_record.csv"
WORKBOOK_TEMPLATE = TEMPLATE_DIR / "ir_swap_revalidation_template.xlsx"


def pick(row: pd.Series, *names: str, default=None):
    for name in names:
        if name in row.index and pd.notna(row[name]):
            return row[name]
    return default


def pick_col(df: pd.DataFrame, *names: str) -> str:
    for name in names:
        if name in df.columns:
            return name
    raise KeyError(f"None of these columns exist: {names}")


def fmt(value: float) -> str:
    return f"{float(value):,.6f}"


def build_excel_bridge_outputs() -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)

    summary = pd.read_csv(SUMMARY_PATH).iloc[0]
    shock = pd.read_csv(SHOCK_PATH)
    lifecycle = pd.read_csv(LIFECYCLE_PATH)

    shock_col = pick_col(shock, "curve_shift_bp", "shock_bp", "parallel_curve_shock_bp", "curve_shock_bp")
    payer_change_col = pick_col(shock, "payer_npv_change", "payer_swap_npv_change", "payer_pnl_change")

    minus100 = shock.loc[shock[shock_col].astype(float).eq(-100.0)].iloc[0]
    plus100 = shock.loc[shock[shock_col].astype(float).eq(100.0)].iloc[0]

    payer_npv = float(pick(summary, "payer_swap_npv", "payer_npv", default=0.0))
    receiver_npv = float(pick(summary, "receiver_swap_npv", "receiver_npv", default=0.0))
    payer_dv01 = float(pick(summary, "payer_dv01", default=0.0))
    receiver_dv01 = float(pick(summary, "receiver_dv01", default=0.0))
    symmetry_error = payer_npv + receiver_npv

    control = pd.DataFrame(
        [
            ["Model ID", "QMRL-IR-SWAP-001", "identifier", "Archer/MRM model record reference"],
            ["Product", "Fixed-for-floating interest-rate swap", "scope", "Vanilla IR derivative"],
            ["Valuation date", str(pick(summary, "valuation_date", default="")), "input", "Official valuation date"],
            ["Par rate", fmt(float(pick(summary, "par_swap_rate", default=0.0))), "input", "Base fair fixed rate"],
            ["Test fixed rate", fmt(float(pick(summary, "validation_fixed_rate", default=0.0))), "input", "Off-market test fixed rate"],
            ["Payer NPV", fmt(payer_npv), "output", "Payer fixed swap value"],
            ["Receiver NPV", fmt(receiver_npv), "output", "Receiver fixed swap value"],
            ["PV symmetry error", fmt(symmetry_error), "control", "Must be close to zero"],
            ["Payer DV01", fmt(payer_dv01), "risk", "Expected positive"],
            ["Receiver DV01", fmt(receiver_dv01), "risk", "Expected negative"],
            ["+100bp payer P&L", fmt(float(plus100[payer_change_col])), "shock", "Expected positive"],
            ["-100bp payer P&L", fmt(float(minus100[payer_change_col])), "shock", "Expected negative"],
            ["XVA coverage", "NOT COVERED", "model boundary", "Requires v0.9 XVA layer"],
            ["Model-use status", "BASE USE ONLY", "decision", "Use for clean pricing and first-order rate-risk checks only"],
        ],
        columns=["field", "value", "type", "decision_meaning"],
    )

    control.to_csv(CONTROL_EXPORT, index=False)
    shock.to_csv(SHOCK_EXPORT, index=False)
    lifecycle.to_csv(LIFECYCLE_EXPORT, index=False)

    workbook_created = create_workbook(control, shock, lifecycle)

    report = f"""# Excel/VBA Revalidation Bridge

## Purpose

This layer connects the Python IR derivatives validation engine to an Excel/VBA review workflow.

Python remains the controlled pricing and validation engine. Excel/VBA provides a reviewer-facing revalidation bridge for teams that still use spreadsheet-based controls.

## Generated outputs

| Artifact | Path |
|---|---|
| Control panel CSV | `{CONTROL_EXPORT.relative_to(ROOT)}` |
| Shock table CSV | `{SHOCK_EXPORT.relative_to(ROOT)}` |
| Lifecycle record CSV | `{LIFECYCLE_EXPORT.relative_to(ROOT)}` |
| VBA module | `excel_vba/vba/Revalidate_IR_Swap.bas` |
| Workbook template | `{WORKBOOK_TEMPLATE.relative_to(ROOT) if workbook_created else "not created because openpyxl is unavailable"}` |

## VBA validation controls

| Control | Rule | Decision use |
|---|---|---|
| PV symmetry | payer NPV plus receiver NPV close to zero | Detect side mismatch |
| DV01 sign | payer DV01 positive, receiver DV01 negative | Detect rate-risk sign error |
| Shock direction | -100bp payer P&L negative and +100bp payer P&L positive | Detect curve-shock inconsistency |
| XVA coverage | marked as NOT COVERED | Prevent overclaiming model coverage |

## Archer/MRM use

The bridge supports a model lifecycle record by exporting model ID, product scope, validation status, monitoring trigger and next lifecycle action.

Generated at: {datetime.now().isoformat(timespec="seconds")}
"""
    REPORT_PATH.write_text(report, encoding="utf-8")

    print("Excel/VBA bridge export complete.")
    print(f"Generated control export: {CONTROL_EXPORT.relative_to(ROOT)}")
    print(f"Generated shock export: {SHOCK_EXPORT.relative_to(ROOT)}")
    print(f"Generated lifecycle export: {LIFECYCLE_EXPORT.relative_to(ROOT)}")
    print(f"Generated report: {REPORT_PATH.relative_to(ROOT)}")
    if workbook_created:
        print(f"Generated workbook template: {WORKBOOK_TEMPLATE.relative_to(ROOT)}")
    else:
        print("Workbook template skipped because openpyxl is unavailable.")


def create_workbook(control: pd.DataFrame, shock: pd.DataFrame, lifecycle: pd.DataFrame) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Control_Panel"
    shock_ws = wb.create_sheet("Shock_Table")
    lifecycle_ws = wb.create_sheet("Lifecycle_Record")
    log_ws = wb.create_sheet("Revalidation_Log")
    inst_ws = wb.create_sheet("Instructions")

    header_fill = PatternFill("solid", fgColor="E8EEF5")
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    thin = Side(style="thin", color="B8C2CE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "IR Swap Excel/VBA Revalidation Bridge"
    ws["A1"].font = title_font

    for col_idx, name in enumerate(control.columns, start=1):
        cell = ws.cell(3, col_idx, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row_idx, record in enumerate(control.itertuples(index=False), start=4):
        for col_idx, value in enumerate(record, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    checks = [
        ["PV symmetry", '=IF(ABS(VALUE(B11))<=1,"PASS","REVIEW")', "payer NPV plus receiver NPV close to zero"],
        ["DV01 sign", '=IF(AND(VALUE(B12)>0,VALUE(B13)<0),"PASS","REVIEW")', "payer DV01 positive and receiver DV01 negative"],
        ["Shock direction", '=IF(AND(VALUE(B16)<0,VALUE(B15)>0),"PASS","REVIEW")', "-100bp negative and +100bp positive"],
        ["Model use", "BASE USE ONLY", "clean pricing and first-order rate risk only"],
        ["XVA", "NOT COVERED", "requires v0.9 XVA layer"],
    ]

    start_row = 20
    ws.cell(start_row, 1, "Excel formula checks").font = header_font
    for col_idx, name in enumerate(["Control", "Status", "Meaning"], start=1):
        cell = ws.cell(start_row + 1, col_idx, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row_idx, record in enumerate(checks, start=start_row + 2):
        for col_idx, value in enumerate(record, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for sheet, frame in [(shock_ws, shock), (lifecycle_ws, lifecycle)]:
        for col_idx, name in enumerate(frame.columns, start=1):
            cell = sheet.cell(1, col_idx, name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        for row_idx, record in enumerate(frame.itertuples(index=False), start=2):
            for col_idx, value in enumerate(record, start=1):
                cell = sheet.cell(row_idx, col_idx, value)
                cell.border = border

    log_ws.append(["timestamp", "user", "status"])
    for col_idx in range(1, 4):
        cell = log_ws.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    inst_ws["A1"] = "Instructions"
    inst_ws["A1"].font = title_font
    inst_ws["A3"] = "Import excel_vba/vba/Revalidate_IR_Swap.bas into this workbook."
    inst_ws["A4"] = "Run RunIRSwapRevalidation() to populate the revalidation status fields."
    inst_ws["A5"] = "Python remains the pricing and validation engine."

    for sheet in [ws, shock_ws, lifecycle_ws, log_ws, inst_ws]:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 24
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(WORKBOOK_TEMPLATE)
    return True


if __name__ == "__main__":
    build_excel_bridge_outputs()
